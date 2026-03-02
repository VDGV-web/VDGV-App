import streamlit as st
import json
import os
import pandas as pd
import re
from datetime import datetime, date
import io
from openpyxl import load_workbook

# ---------------- Pfade ----------------
BASE_DIR = r"D:\App_DGM\daten"

FAHRER_DATEI = os.path.join(BASE_DIR, "nennungen_fahrer.json")
MANNSCHAFT_DATEI = os.path.join(BASE_DIR, "nennungen_mannschaft.json")
TERMINE_DATEI = os.path.join(BASE_DIR, "termine.json")
FAHRZEUGE_CSV = os.path.join(BASE_DIR, "fahrzeuge.csv")
KLASSEN_CSV = os.path.join(BASE_DIR, "klassen.csv")
MANNSCHAFTEN_CSV = os.path.join(BASE_DIR, "mannschaften.csv")
STAMMDATEN_FAHRZEUG = os.path.join(BASE_DIR, "FahrzeugAbnahme_Stammdaten.xlsx")
HCF_EXCEL = os.path.join(BASE_DIR, "HCF.xlsx")
BOARDKARTE_TEMPLATE = os.path.join(BASE_DIR, "Boardkarte_A4_Landscape_Final.xlsx")


# ---------------- Klassenfarben ----------------
KLASSEN_REIHENFOLGE = [
    "junior-cup","fun-cup","offene klasse","original",
    "standard","modified","promodified","prototype"
]

KLASSEN_FARBEN = {
    "fun-cup": "#4CAF50",
    "junior-cup": "#9C27B0",
    "original": "#2196F3",
    "standard": "#FFFFFF",
    "modified": "#FFEB3B",
    "promodified": "#000000",
    "prototype": "#F44336",
    "offene klasse": "#FF9800",
}


# ---------------- Helper ----------------
def lade_json(datei):
    if os.path.exists(datei):
        try:
            with open(datei,"r",encoding="utf-8") as f:
                d=json.load(f)
                return d if isinstance(d,list) else []
        except:
            return []
    return []


def speichere_json(datei,daten):
    with open(datei,"w",encoding="utf-8") as f:
        json.dump(daten,f,indent=4,ensure_ascii=False)


# ---------------- CSV Loader (1 Spalte sicher) ----------------
def lade_csv(datei, spalten, fallback=[]):
    if not os.path.exists(datei):
        return fallback

    try:
        df = pd.read_csv(datei, header=None, encoding="utf-8").fillna("")
    except:
        return fallback

    first = str(df.iloc[0,0]).lower().strip()

    # falls Header in 1. Zeile
    if first in [s.lower() for s in spalten]:
        df.columns = [df.iloc[0,0]]
        df = df.iloc[1:]

    # 1-Spalten CSV
    if len(df.columns) == 1:
        return sorted(df.iloc[:,0].astype(str).str.strip().unique())

    # Multi-Spalten: Spaltenname suchen
    for s in spalten:
        for c in df.columns:
            if c.lower()==s.lower():
                return sorted(df[c].astype(str).str.strip().unique())

    return fallback


# ---------------- Stammdaten + Abnahme ----------------
def normalize_number(s):
    if not s: return None
    d = re.sub(r"\D", "", str(s))
    return int(d) if d else None


def lade_stammdaten():
    if not os.path.exists(STAMMDATEN_FAHRZEUG):
        return pd.DataFrame(columns=["Startnummer","Abnahme Datum"])

    df = pd.read_excel(STAMMDATEN_FAHRZEUG, dtype=str).fillna("")
    df.columns = df.columns.str.strip()

    if "Abnahme Datum" not in df.columns:
        df["Abnahme Datum"] = ""

    df["Abnahme Datum"] = pd.to_datetime(df["Abnahme Datum"], errors="coerce")
    return df


def find_abnahme(stammdaten, startnummer):
    sn = normalize_number(startnummer)
    hit = stammdaten.loc[stammdaten["Startnummer"].apply(normalize_number)==sn]
    if hit.empty:
        return False, None

    dat = hit.iloc[0]["Abnahme Datum"]
    if pd.isna(dat):
        return False, None

    return True, dat.date()


# ---------------- HCF ----------------
def lade_hcf():
    if not os.path.exists(HCF_EXCEL):
        return pd.DataFrame(columns=["auto","auto_norm","hcf"])

    df=pd.read_excel(HCF_EXCEL,dtype=str).fillna("")
    df.columns=[c.lower().strip() for c in df.columns]

    auto_col=next((c for c in df.columns if "auto" in c),None)
    hcf_col=next((c for c in df.columns if "hcf" in c),None)

    if not auto_col or not hcf_col:
        return pd.DataFrame(columns=["auto","auto_norm","hcf"])

    df["auto"]=df[auto_col].astype(str)
    df["auto_norm"]=df["auto"].str.lower()
    df["hcf"]=df[hcf_col]
    return df


def match_hcf(auto,df):
    if not auto or df.empty: return ""
    a=str(auto).lower()
    hit=df[df["auto_norm"]==a]
    if not hit.empty:
        return hit.iloc[0]["hcf"]
    return ""


# ---------------- BOARDKARTE ----------------
def tmpl():
    if os.path.exists(BOARDKARTE_TEMPLATE):
        return load_workbook(BOARDKARTE_TEMPLATE)
    return None


def fuelle_boardkarte(fahrer,hcf):
    wb=tmpl()
    if wb is None:
        out=io.BytesIO()
        return out
    ws=wb.active

    ph = {
        "{{Fahrzeug}}": fahrer.get("auto",""),
        "{{Startnummer}}": fahrer.get("startnummer",""),
        "{{Klasse}}": fahrer.get("klasse",""),
        "{{Fahrer}}": f"{fahrer.get('vorname','')} {fahrer.get('name','')}",
        "{{Beifahrer}}": fahrer.get("beifahrer",""),
        "{{HCF}}": hcf,
        "{{Abnahme}}": fahrer.get("abnahme_datum","")
    }

    for r in ws.iter_rows():
        for cell in r:
            if isinstance(cell.value,str):
                v=cell.value
                for k,val in ph.items():
                    v=v.replace(k,str(val))
                cell.value=v

    out=io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out


# ---------------- UI ----------------
def show():
    st.set_page_config(layout="wide")
    st.title("🏎 Nennbüro – MIT TECHNISCHER ABNAHME")

    stammdaten = lade_stammdaten()
    hcf_tab = lade_hcf()

    klassen = lade_csv(KLASSEN_CSV,["klasse","klassen"],[])
    autos = lade_csv(FAHRZEUGE_CSV,["fahrzeug","auto"],[])
    vereine = lade_csv(MANNSCHAFTEN_CSV,["verein","club"],[])

    fahrer = lade_json(FAHRER_DATEI)
    termine_raw = lade_json(TERMINE_DATEI)

    termine = []
    for t in termine_raw:
        if isinstance(t,dict) and "datum" in t:
            termine.append(f"{t['datum']} – {t.get('beschreibung','')}")

    if not termine:
        st.warning("Keine Termine!")
        return

    tabs=st.tabs(termine)

    # =====================================
    for tidx,termin in enumerate(termine):
        with tabs[tidx]:

            subset=[f for f in fahrer if f.get("lauf")==termin]
            subset.sort(key=lambda f:(
                KLASSEN_REIHENFOLGE.index(str(f.get("klasse","")).lower())
                if str(f.get("klasse","")).lower() in KLASSEN_REIHENFOLGE else 999,
                normalize_number(f.get("startnummer")) or 99999
            ))

            cur_class=None

            for i,f in enumerate(subset):
                klasse=str(f.get("klasse","")).lower()

                # farbstreifen
                if klasse!=cur_class:
                    cur_class=klasse
                    col=KLASSEN_FARBEN.get(klasse,"#ccc")
                    st.markdown(
                        f"<div style='background:{col};padding:6px;border-radius:4px;margin:8px 0;color:#fff'>{klasse.upper()}</div>",
                        unsafe_allow_html=True
                    )

                # automatische Abnahme aus Stammdaten
                auto_ok, auto_date = find_abnahme(stammdaten, f.get("startnummer"))
                if auto_ok and not f.get("abnahme_status"):
                    f["abnahme_status"]=True
                    f["abnahme_datum"]=auto_date.strftime("%Y-%m-%d")

                hcf_val=match_hcf(f.get("auto"),hcf_tab)

                with st.expander(f"{f.get('vorname')} {f.get('name')} — #{f.get('startnummer')}", expanded=False):

                    # Fahrer UI
                    c1,c2,c3 = st.columns(3)
                    f["vorname"]=c1.text_input("Vorname",f.get("vorname",""),key=f"v{tidx}{i}")
                    f["name"]=c2.text_input("Nachname",f.get("name",""),key=f"n{tidx}{i}")
                    f["klasse"]=c3.selectbox("Klasse",klassen,index=klassen.index(f["klasse"]) if f.get("klasse") in klassen else 0,key=f"k{tidx}{i}")

                    d1,d2,d3 = st.columns(3)
                    f["startnummer"]=d1.text_input("Startnummer",f.get("startnummer",""),key=f"s{tidx}{i}")
                    f["auto"]=d2.selectbox("Auto",autos,index=autos.index(f["auto"]) if f.get("auto") in autos else 0,key=f"a{tidx}{i}")
                    f["verein"]=d3.selectbox("Verein",vereine,index=vereine.index(f["verein"]) if f.get("verein") in vereine else 0,key=f"g{tidx}{i}")

                    f["beifahrer"]=st.text_input("Beifahrer",f.get("beifahrer",""),key=f"b{tidx}{i}")

                    # --- Abnahme ---
                    st.subheader("🛠 Technische Abnahme")
                    f["abnahme_status"]=st.checkbox("Abnahme durchgeführt",value=f.get("abnahme_status",False),key=f"ab{tidx}{i}")

                    if f["abnahme_status"]:

                        raw_date = f.get("abnahme_datum", auto_date)

                        # convert sicher
                        try:
                            if isinstance(raw_date,str):
                                init_date = pd.to_datetime(raw_date).date()
                            else:
                                init_date = raw_date or datetime.now().date()
                        except:
                            init_date = datetime.now().date()

                        f["abnahme_datum"]=st.date_input(
                            "Datum der Abnahme",
                            value=init_date,
                            key=f"ad{tidx}{i}"
                        ).strftime("%Y-%m-%d")

                        f["abnahme_kommentar"]=st.text_area(
                            "Kommentar zur Abnahme",
                            value=f.get("abnahme_kommentar",""),
                            key=f"ak{tidx}{i}"
                        )
                    else:
                        f["abnahme_datum"]=""
                        f["abnahme_kommentar"]=""

                    st.info(f"📦 **HCF**: {hcf_val}")

                    bezahlt=st.checkbox("💰 bezahlt",value=f.get("bezahlt",False),key=f"p{tidx}{i}")
                    f["bezahlt"]=bool(bezahlt)

                    if st.button("Speichern",key=f"z{tidx}{i}"):
                        speichere_json(FAHRER_DATEI,fahrer)
                        st.success("Gespeichert ✔️")

                    dl=fuelle_boardkarte(f,hcf_val)
                    st.download_button(
                        "📄 Boardkarte",
                        dl,
                        file_name=f"Boardkarte_{f.get('startnummer')}_{f.get('name')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key=f"x{tidx}{i}"
                    )


if __name__ == "__main__":
    show()


